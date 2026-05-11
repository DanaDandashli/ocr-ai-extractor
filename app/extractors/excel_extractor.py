from openpyxl import load_workbook


"""
    This function extract text from Excel (.xlsx) files.
"""
def extract_text_from_excel(file_path: str) -> str:
    workbook = load_workbook(file_path, data_only=True)
    text_parts = []

    for sheet in workbook.worksheets:
        text_parts.append(f"Sheet: {sheet.title}")

        for row in sheet.iter_rows(values_only=True):
            row_data = []

            for cell in row:
                if cell is not None and str(cell).strip():
                    row_data.append(str(cell).strip())

            if row_data:
                text_parts.append(" | ".join(row_data))

    return "\n".join(text_parts).strip()
